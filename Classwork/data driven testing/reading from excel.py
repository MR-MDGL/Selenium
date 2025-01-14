import os
import openpyxl
#file-----worknook-----sheet-----rows---excel
file=os.getcwd() +r'\testdata.xlsx'

workbook=openpyxl.load_workbook(file)
sheet=workbook['Sheet1']
rows=sheet.max_row
print('totlal rows',rows)
cols=sheet.max_column
print('totlal cols',cols)

# reading alll the rows and cal from excel
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="   ")
    print()


